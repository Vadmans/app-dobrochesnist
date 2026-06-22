"""
Код Доброчесності — бекенд + API + адмін-панель.
Виправлено:
- події + Excel імпорт у вкладці Події;
- прибрано верхні кнопки '+ Нова подія' і '/docs';
- додано вкладку Довідник;
- статистика не падає, навіть якщо частина таблиць порожня;
- push завжди повертає JSON з реальною причиною помилки;
- сумісність із мобільним App.jsx: /chat/question і /chat/messages.
Excel: A = дата, B = подія, C = посилання.
"""
import os
import uuid
import secrets
import json
from datetime import date, datetime, timedelta
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy import create_engine, update, String, Integer, Date, DateTime, JSON, Text, text
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, Session, sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./dobrochesnist.db")
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, connect_args=connect_args, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)

class Base(DeclarativeBase):
    pass

class Event(Base):
    __tablename__ = "events"
    id: Mapped[str] = mapped_column(String, primary_key=True)
    cat: Mapped[str] = mapped_column(String, index=True, default="notice")
    title: Mapped[str] = mapped_column(String)
    date: Mapped[date] = mapped_column(Date, index=True)
    recur: Mapped[str] = mapped_column(String, default="")
    description: Mapped[str] = mapped_column(Text, default="")
    instruction: Mapped[str] = mapped_column(Text, default="")
    link: Mapped[str] = mapped_column(Text, default="")
    audience: Mapped[str] = mapped_column(String, default="Усі працівники")
    reminders: Mapped[list] = mapped_column(JSON, default=list)
    views: Mapped[int] = mapped_column(Integer, default=0)

class ReferenceItem(Base):
    __tablename__ = "reference_items"
    id: Mapped[str] = mapped_column(String, primary_key=True)
    title: Mapped[str] = mapped_column(String)
    body: Mapped[str] = mapped_column(Text, default="")
    link: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

class Device(Base):
    __tablename__ = "devices"
    id: Mapped[str] = mapped_column(String, primary_key=True)
    token: Mapped[str] = mapped_column(Text, unique=True, index=True)
    platform: Mapped[str] = mapped_column(String, default="android")
    app_version: Mapped[str] = mapped_column(String, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

class ChatMessage(Base):
    __tablename__ = "chat_messages"
    id: Mapped[str] = mapped_column(String, primary_key=True)
    client_id: Mapped[str] = mapped_column(String, index=True)
    question: Mapped[str] = mapped_column(Text)
    answer: Mapped[str] = mapped_column(Text, default="")
    status: Mapped[str] = mapped_column(String, default="new")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    answered_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)

class PushLog(Base):
    __tablename__ = "push_logs"
    id: Mapped[str] = mapped_column(String, primary_key=True)
    title: Mapped[str] = mapped_column(String)
    body: Mapped[str] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    sent: Mapped[int] = mapped_column(Integer, default=0)
    failed: Mapped[int] = mapped_column(Integer, default=0)
    configured: Mapped[int] = mapped_column(Integer, default=0)
    error: Mapped[str] = mapped_column(Text, default="")

Base.metadata.create_all(engine)

def ensure_schema():
    """Мінімальна міграція для старої БД без Alembic.
    Важливо для Render/PostgreSQL: create_all НЕ додає колонки в існуючі таблиці.
    """
    try:
        with engine.begin() as conn:
            if DATABASE_URL.startswith("sqlite"):
                def sqlite_cols(table):
                    try:
                        return [r[1] for r in conn.execute(text(f"PRAGMA table_info({table})"))]
                    except Exception:
                        return []
                cols = sqlite_cols("events")
                if "link" not in cols:
                    conn.execute(text("ALTER TABLE events ADD COLUMN link TEXT DEFAULT ''"))
                dcols = sqlite_cols("devices")
                if dcols:
                    for col, ddl in [
                        ("id", "TEXT"), ("platform", "TEXT DEFAULT 'android'"),
                        ("app_version", "TEXT DEFAULT ''"), ("created_at", "DATETIME"),
                        ("updated_at", "DATETIME")
                    ]:
                        if col not in dcols:
                            conn.execute(text(f"ALTER TABLE devices ADD COLUMN {col} {ddl}"))
                pcols = sqlite_cols("push_logs")
                if "error" not in pcols and pcols:
                    conn.execute(text("ALTER TABLE push_logs ADD COLUMN error TEXT DEFAULT ''"))
            else:
                conn.execute(text("ALTER TABLE events ADD COLUMN IF NOT EXISTS link TEXT DEFAULT ''"))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS devices (
                        token TEXT PRIMARY KEY,
                        platform VARCHAR DEFAULT 'android',
                        app_version VARCHAR DEFAULT '',
                        created_at TIMESTAMP DEFAULT NOW(),
                        updated_at TIMESTAMP DEFAULT NOW()
                    )
                """))
                conn.execute(text("ALTER TABLE devices ADD COLUMN IF NOT EXISTS id TEXT"))
                conn.execute(text("ALTER TABLE devices ADD COLUMN IF NOT EXISTS platform VARCHAR DEFAULT 'android'"))
                conn.execute(text("ALTER TABLE devices ADD COLUMN IF NOT EXISTS app_version VARCHAR DEFAULT ''"))
                conn.execute(text("ALTER TABLE devices ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW()"))
                conn.execute(text("ALTER TABLE devices ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()"))
                conn.execute(text("UPDATE devices SET id = 'd' || substr(md5(token), 1, 10) WHERE id IS NULL OR id = ''"))
                conn.execute(text("ALTER TABLE push_logs ADD COLUMN IF NOT EXISTS error TEXT DEFAULT ''"))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS reference_items (
                        id VARCHAR PRIMARY KEY,
                        title VARCHAR NOT NULL,
                        body TEXT DEFAULT '',
                        link TEXT DEFAULT '',
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                """))
    except Exception as e:
        print("ensure_schema warning:", e)
ensure_schema()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

ADMIN_TOKEN = os.getenv("ADMIN_TOKEN")
if not ADMIN_TOKEN:
    ADMIN_TOKEN = secrets.token_urlsafe(24)
    print("\n" + "="*64)
    print("ADMIN_TOKEN не задано — тимчасовий токен:")
    print(ADMIN_TOKEN)
    print("Для Render задай ADMIN_TOKEN у Environment Variables")
    print("="*64 + "\n")

def require_admin(authorization: str = Header(default="")):
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not secrets.compare_digest(token, ADMIN_TOKEN):
        raise HTTPException(401, "Потрібна авторизація адміністратора", headers={"WWW-Authenticate":"Bearer"})

class EventIn(BaseModel):
    cat: str = "notice"
    title: str
    date: date
    recur: str = ""
    description: str = ""
    instruction: str = ""
    link: str = ""
    audience: str = "Усі працівники"
    reminders: List[int] = []

class EventOut(EventIn):
    id: str
    views: int
    class Config:
        from_attributes = True

class ReferenceIn(BaseModel):
    title: str
    body: str = ""
    link: str = ""

class ReferenceOut(ReferenceIn):
    id: str
    created_at: datetime
    class Config:
        from_attributes = True

class DeviceIn(BaseModel):
    token: str
    platform: str = "android"
    app_version: str = ""

class ChatIn(BaseModel):
    client_id: str
    question: str

class AnswerIn(BaseModel):
    answer: str

class PushIn(BaseModel):
    title: str
    body: str

app = FastAPI(title="Код Доброчесності API", version="0.4.0")
_origins_env = os.getenv("ALLOWED_ORIGINS", "*").strip()
allow_origins = ["*"] if _origins_env == "*" else [o.strip() for o in _origins_env.split(",") if o.strip()]
app.add_middleware(CORSMiddleware, allow_origins=allow_origins, allow_methods=["*"], allow_headers=["*"])

@app.exception_handler(Exception)
async def json_exception_handler(request, exc):
    print("Unhandled error:", repr(exc))
    return JSONResponse(status_code=500, content={"ok": False, "detail": "Internal Server Error", "error": str(exc)})

# ---------- Події ----------
@app.get("/events", response_model=List[EventOut])
def list_events(db: Session = Depends(get_db)):
    return db.query(Event).order_by(Event.date).all()

@app.get("/events/{event_id}", response_model=EventOut)
def get_event(event_id: str, db: Session = Depends(get_db)):
    ev = db.get(Event, event_id)
    if not ev:
        raise HTTPException(404, "Подію не знайдено")
    return ev

@app.post("/events/{event_id}/view", response_model=EventOut)
def register_view(event_id: str, db: Session = Depends(get_db)):
    ev = db.get(Event, event_id)
    if not ev:
        raise HTTPException(404, "Подію не знайдено")
    db.execute(update(Event).where(Event.id == event_id).values(views=Event.views + 1))
    db.commit(); db.refresh(ev)
    return ev

@app.post("/events", response_model=EventOut, status_code=201, dependencies=[Depends(require_admin)])
def create_event(data: EventIn, db: Session = Depends(get_db)):
    ev = Event(id=f"e{uuid.uuid4().hex[:8]}", views=0, **data.model_dump())
    db.add(ev); db.commit(); db.refresh(ev)
    return ev

@app.put("/events/{event_id}", response_model=EventOut, dependencies=[Depends(require_admin)])
def update_event(event_id: str, data: EventIn, db: Session = Depends(get_db)):
    ev = db.get(Event, event_id)
    if not ev:
        raise HTTPException(404, "Подію не знайдено")
    for k, v in data.model_dump().items():
        setattr(ev, k, v)
    db.commit(); db.refresh(ev)
    return ev

@app.delete("/events/{event_id}", status_code=204, dependencies=[Depends(require_admin)])
def delete_event(event_id: str, db: Session = Depends(get_db)):
    ev = db.get(Event, event_id)
    if not ev:
        raise HTTPException(404, "Подію не знайдено")
    db.delete(ev); db.commit()

def parse_excel_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).date()
        except Exception:
            return None
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

@app.post("/events/import-excel", dependencies=[Depends(require_admin)])
async def import_events_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Завантажте файл Excel .xlsx")
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, "На сервері не встановлено openpyxl. Додайте openpyxl у requirements.txt")
    try:
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Не вдалося прочитати Excel: {e}")
    created, skipped = 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        raw_date = row[0] if len(row) > 0 else None
        raw_title = row[1] if len(row) > 1 else None
        raw_link = row[2] if len(row) > 2 else ""
        ev_date = parse_excel_date(raw_date)
        title = str(raw_title or "").strip()
        link = str(raw_link or "").strip()
        if idx == 1 and str(raw_date or "").strip().lower() in ("дата", "date"):
            continue
        if not ev_date or not title:
            if raw_date or raw_title or raw_link:
                skipped.append({"row": idx, "reason": "немає дати або назви"})
            continue
        # Важливо: cat="notice", бо мобільний додаток краще працює з існуючими категоріями.
        ev = Event(
            id=f"e{uuid.uuid4().hex[:8]}", cat="notice", title=title, date=ev_date,
            recur="", description=(f"Посилання: {link}" if link else ""), instruction=link,
            link=link, audience="Усі працівники", reminders=[30, 10, 3, 0], views=0
        )
        db.add(ev); created += 1
    db.commit()
    return {"ok": True, "created": created, "skipped": skipped[:50], "skipped_count": len(skipped)}

# ---------- Довідник ----------
def table_exists(db: Session, name: str) -> bool:
    try:
        if DATABASE_URL.startswith("sqlite"):
            return db.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name=:n"), {"n": name}).first() is not None
        return db.execute(text("SELECT to_regclass(:n)"), {"n": name}).scalar() is not None
    except Exception:
        return False

def row_to_reference_dict(row, fallback_id_prefix="r"):
    m = dict(row._mapping) if hasattr(row, "_mapping") else dict(row)
    rid = str(m.get("id") or m.get("uuid") or m.get("reference_id") or f"{fallback_id_prefix}{uuid.uuid4().hex[:8]}")
    title = str(m.get("title") or m.get("name") or m.get("question") or m.get("caption") or "Без назви")
    body = str(m.get("body") or m.get("description") or m.get("text") or m.get("content") or m.get("answer") or "")
    link = str(m.get("link") or m.get("url") or "")
    created = m.get("created_at") or m.get("date") or datetime.utcnow()
    if isinstance(created, str):
        try:
            created = datetime.fromisoformat(created.replace("Z", "+00:00"))
        except Exception:
            created = datetime.utcnow()
    created_str = created.isoformat() if hasattr(created, "isoformat") else str(created)
    return {"id": rid, "title": title, "body": body, "description": body, "link": link, "created_at": created_str}

@app.get("/reference")
def list_reference(db: Session = Depends(get_db)):
    items = []
    try:
        for x in db.query(ReferenceItem).order_by(ReferenceItem.created_at.desc()).all():
            items.append({"id": x.id, "title": x.title, "body": x.body, "description": x.body, "link": x.link, "created_at": x.created_at.isoformat()})
    except Exception:
        pass
    for tname in ("referens", "reference", "references"):
        if table_exists(db, tname):
            try:
                rows = db.execute(text(f"SELECT * FROM {tname} LIMIT 500")).fetchall()
                for r in rows:
                    d = row_to_reference_dict(r, tname[0])
                    if not any(x["id"] == d["id"] and x["title"] == d["title"] for x in items):
                        items.append(d)
            except Exception as e:
                print(f"reference compatibility warning {tname}:", e)
    return items

@app.post("/reference", status_code=201, dependencies=[Depends(require_admin)])
def create_reference(data: ReferenceIn, db: Session = Depends(get_db)):
    item = ReferenceItem(id=f"r{uuid.uuid4().hex[:8]}", **data.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return {"id": item.id, "title": item.title, "body": item.body, "description": item.body, "link": item.link, "created_at": item.created_at.isoformat()}

@app.put("/reference/{item_id}", dependencies=[Depends(require_admin)])
def update_reference(item_id: str, data: ReferenceIn, db: Session = Depends(get_db)):
    item = db.get(ReferenceItem, item_id)
    if not item:
        raise HTTPException(404, "Запис не знайдено")
    item.title = data.title; item.body = data.body; item.link = data.link
    db.commit(); db.refresh(item)
    return {"id": item.id, "title": item.title, "body": item.body, "description": item.body, "link": item.link, "created_at": item.created_at.isoformat()}

@app.delete("/reference/{item_id}", status_code=204, dependencies=[Depends(require_admin)])
def delete_reference(item_id: str, db: Session = Depends(get_db)):
    item = db.get(ReferenceItem, item_id)
    if not item:
        raise HTTPException(404, "Запис не знайдено")
    db.delete(item); db.commit()

# ---------- Пристрої / Push ----------
def fetch_device_rows(db: Session):
    try:
        if DATABASE_URL.startswith("sqlite"):
            q = "SELECT token, COALESCE(id, '') AS id, COALESCE(platform, 'android') AS platform, COALESCE(app_version, '') AS app_version, COALESCE(updated_at, created_at) AS updated_at FROM devices WHERE token IS NOT NULL AND token <> ''"
        else:
            q = "SELECT token, COALESCE(id, '') AS id, COALESCE(platform, 'android') AS platform, COALESCE(app_version, '') AS app_version, COALESCE(updated_at, created_at, NOW()) AS updated_at FROM devices WHERE token IS NOT NULL AND token <> ''"
        rows = db.execute(text(q)).fetchall()
        return [dict(r._mapping) for r in rows]
    except Exception as e:
        print("fetch_device_rows warning:", e)
        return []

@app.post("/devices/register")
def register_device(data: DeviceIn, db: Session = Depends(get_db)):
    token = data.token.strip()
    if not token:
        raise HTTPException(400, "Порожній token")
    did = f"d{uuid.uuid4().hex[:10]}"
    try:
        if DATABASE_URL.startswith("sqlite"):
            db.execute(text("""
                INSERT INTO devices (id, token, platform, app_version, created_at, updated_at)
                VALUES (:id, :token, :platform, :app_version, :now, :now)
                ON CONFLICT(token) DO UPDATE SET platform=:platform, app_version=:app_version, updated_at=:now
            """), {"id": did, "token": token, "platform": data.platform, "app_version": data.app_version, "now": datetime.utcnow()})
        else:
            db.execute(text("""
                INSERT INTO devices (id, token, platform, app_version, created_at, updated_at)
                VALUES (:id, :token, :platform, :app_version, NOW(), NOW())
                ON CONFLICT(token) DO UPDATE SET platform=EXCLUDED.platform, app_version=EXCLUDED.app_version, updated_at=NOW(), id=COALESCE(devices.id, EXCLUDED.id)
            """), {"id": did, "token": token, "platform": data.platform, "app_version": data.app_version})
        db.commit()
        return {"ok": True, "device_id": did}
    except Exception as e:
        db.rollback()
        return JSONResponse(status_code=200, content={"ok": False, "error": str(e)})

@app.get("/devices", dependencies=[Depends(require_admin)])
def list_devices(db: Session = Depends(get_db)):
    items = fetch_device_rows(db)[:200]
    return [{"id": d.get("id") or "", "platform": d.get("platform") or "", "app_version": d.get("app_version") or "", "updated_at": str(d.get("updated_at") or ""), "token_start": (d.get("token") or "")[:18]} for d in items]

def send_push_to_tokens(tokens: List[str], title: str, body: str):
    service_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
    if not service_json:
        return {"configured": False, "sent": 0, "failed": 0, "error": "FIREBASE_SERVICE_ACCOUNT_JSON не задано в Render Environment"}
    if not tokens:
        return {"configured": True, "sent": 0, "failed": 0, "error": "Немає зареєстрованих пристроїв. Відкрий додаток на телефоні, щоб він передав FCM token."}
    try:
        import firebase_admin
        from firebase_admin import credentials, messaging
        try:
            service_data = json.loads(service_json)
        except Exception as e:
            return {"configured": False, "sent": 0, "failed": len(tokens), "error": f"FIREBASE_SERVICE_ACCOUNT_JSON не є правильним JSON: {e}"}
        if not firebase_admin._apps:
            cred = credentials.Certificate(service_data)
            firebase_admin.initialize_app(cred)
        sent = failed = 0
        errors = []
        for token in tokens:
            try:
                messaging.send(messaging.Message(notification=messaging.Notification(title=title, body=body), token=token))
                sent += 1
            except Exception as e:
                failed += 1
                if len(errors) < 3:
                    errors.append(str(e))
        return {"configured": True, "sent": sent, "failed": failed, "error": "; ".join(errors)}
    except Exception as e:
        return {"configured": False, "sent": 0, "failed": len(tokens), "error": str(e)}

@app.post("/push/send", dependencies=[Depends(require_admin)])
def push_send(data: PushIn, db: Session = Depends(get_db)):
    try:
        devices = fetch_device_rows(db)
        tokens = [d.get("token") for d in devices if d.get("token")]
        result = send_push_to_tokens(tokens, data.title, data.body)
        log = PushLog(
            id=f"p{uuid.uuid4().hex[:8]}", title=data.title, body=data.body,
            sent=result.get("sent", 0), failed=result.get("failed", 0),
            configured=1 if result.get("configured") else 0, error=result.get("error", "") or ""
        )
        db.add(log); db.commit()
        result["ok"] = bool(result.get("configured"))
        result["devices"] = len(tokens)
        return result
    except Exception as e:
        return JSONResponse(status_code=200, content={"ok": False, "configured": False, "sent": 0, "failed": 0, "devices": 0, "error": str(e)})

@app.get("/push/status", dependencies=[Depends(require_admin)])
def push_status(db: Session = Depends(get_db)):
    return {
        "firebase_env": bool(os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()),
        "devices": len(fetch_device_rows(db)),
        "logs": db.query(PushLog).count(),
    }

# ---------- Чат ----------
@app.post("/chat")
def chat_create(data: ChatIn, db: Session = Depends(get_db)):
    question = data.question.strip()
    if not question:
        raise HTTPException(400, "Порожнє питання")
    msg = ChatMessage(id=f"c{uuid.uuid4().hex[:10]}", client_id=data.client_id.strip() or "unknown", question=question)
    db.add(msg); db.commit()
    return {"ok": True, "id": msg.id}

# Сумісність із мобільним додатком.
@app.post("/chat/question")
def chat_question(data: ChatIn, db: Session = Depends(get_db)):
    return chat_create(data, db)

@app.get("/chat/messages")
def chat_messages(client_id: str, db: Session = Depends(get_db)):
    return chat_for_client(client_id, db)

@app.get("/chat/{client_id}")
def chat_for_client(client_id: str, db: Session = Depends(get_db)):
    items = db.query(ChatMessage).filter(ChatMessage.client_id == client_id).order_by(ChatMessage.created_at.desc()).limit(50).all()
    return [{"id":m.id,"question":m.question,"answer":m.answer,"status":m.status,"created_at":m.created_at.isoformat(),"answered_at":m.answered_at.isoformat() if m.answered_at else None} for m in items]

@app.get("/admin/chat", dependencies=[Depends(require_admin)])
def admin_chat(db: Session = Depends(get_db)):
    items = db.query(ChatMessage).order_by(ChatMessage.created_at.desc()).limit(200).all()
    return [{"id":m.id,"client_id":m.client_id,"question":m.question,"answer":m.answer,"status":m.status,"created_at":m.created_at.isoformat(),"answered_at":m.answered_at.isoformat() if m.answered_at else None} for m in items]

@app.post("/admin/chat/{message_id}/answer", dependencies=[Depends(require_admin)])
def answer_chat(message_id: str, data: AnswerIn, db: Session = Depends(get_db)):
    msg = db.get(ChatMessage, message_id)
    if not msg:
        raise HTTPException(404, "Повідомлення не знайдено")
    msg.answer = data.answer.strip()
    msg.status = "answered" if msg.answer else "new"
    msg.answered_at = datetime.utcnow() if msg.answer else None
    db.commit()
    return {"ok": True}

@app.delete("/admin/chat/{message_id}", status_code=204, dependencies=[Depends(require_admin)])
def delete_chat(message_id: str, db: Session = Depends(get_db)):
    msg = db.get(ChatMessage, message_id)
    if not msg:
        raise HTTPException(404, "Повідомлення не знайдено")
    db.delete(msg); db.commit()

# ---------- Статистика ----------
@app.get("/stats", dependencies=[Depends(require_admin)])
def stats(db: Session = Depends(get_db)):
    try:
        events = db.query(Event).all()
        chat_new = db.query(ChatMessage).filter(ChatMessage.status == "new").count()
        devices = len(fetch_device_rows(db))
        refs = len(list_reference(db))
        try:
            push_logs = db.query(PushLog).count()
        except Exception:
            push_logs = 0
        return {
            "ok": True,
            "events": len(events),
            "views": sum((e.views or 0) for e in events),
            "categories": len({e.cat for e in events}),
            "devices": devices,
            "chat_new": chat_new,
            "reference": refs,
            "push_logs": push_logs,
            "by_event": [{"id": e.id, "title": e.title, "cat": e.cat, "views": e.views or 0} for e in sorted(events, key=lambda e: e.views or 0, reverse=True)],
        }
    except Exception as e:
        return JSONResponse(status_code=200, content={"ok": False, "error": str(e), "events": 0, "views": 0, "categories": 0, "devices": 0, "chat_new": 0, "reference": 0, "push_logs": 0, "by_event": []})

def seed():
    db = SessionLocal()
    try:
        if db.query(Event).count() == 0:
            today = date.today()
            db.add_all([
                Event(id="e1", cat="declaration", title="Подання щорічної декларації", date=today+timedelta(days=12), recur="Щороку, до 1 квітня", description="Подати щорічну декларацію через Реєстр декларацій.", instruction="Перевірте доходи, майно, транспорт та корпоративні права.", link="", audience="Усі працівники", reminders=[30,10,3,0], views=184),
                Event(id="e2", cat="training", title="Щорічне навчання з доброчесності", date=today+timedelta(days=3), recur="Щороку", description="Пройти онлайн-курс із запобігання конфлікту інтересів.", instruction="Після проходження збережіть сертифікат.", link="", audience="Усі працівники", reminders=[10,3,0], views=97),
            ])
        if db.query(ReferenceItem).count() == 0:
            db.add(ReferenceItem(id="r1", title="Довідник доброчесності", body="Тут можна додавати довідкові матеріали для користувачів.", link=""))
        db.commit()
    finally:
        db.close()
seed()

@app.get("/", response_class=HTMLResponse)
def root():
    return """
<!doctype html><html lang="uk"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Код Доброчесності</title>
<style>:root{--n:#102f44;--b:#1f5673;--g:#f2b134;--bg:#f4f7fb;--i:#17212f;--m:#667085}*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:radial-gradient(circle at top left,#deedf5,#f7f9fc 45%,#edf3f8);color:var(--i)}.hero{max-width:1000px;margin:auto;min-height:100vh;display:grid;place-items:center;padding:28px}.card{width:100%;background:rgba(255,255,255,.88);border:1px solid #fff;border-radius:30px;box-shadow:0 24px 70px rgba(16,47,68,.14);padding:42px}.brand{display:flex;gap:14px;align-items:center;margin-bottom:30px}.logo{width:58px;height:58px;border-radius:18px;background:linear-gradient(145deg,var(--n),var(--b));display:grid;place-items:center;color:#fff;font-weight:950;font-size:22px}.brand b{font-size:18px}.brand span{display:block;color:var(--m);font-size:13px}h1{font-size:46px;line-height:1.04;margin:0 0 16px;color:var(--n)}.lead{font-size:17px;line-height:1.7;color:#475467;margin:0 0 26px}.btn{display:inline-block;padding:13px 18px;border-radius:15px;text-decoration:none;font-weight:850;background:var(--b);color:white}@media(max-width:700px){h1{font-size:34px}.card{padding:24px}}</style></head>
<body><main class="hero"><section class="card"><div class="brand"><div class="logo">КД</div><div><b>Код Доброчесності</b><span>Твій цифровий орієнтир</span></div></div><h1>Сервіс працює</h1><p class="lead">Адмін-панель керує подіями, Excel-імпортом, довідником, чатом, статистикою та push-повідомленнями.</p><a class="btn" href="/admin">Відкрити адмін-панель</a></section></main></body></html>
"""

@app.get("/admin", response_class=HTMLResponse)
def admin_panel():
    return r'''
<!doctype html><html lang="uk"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Адмін-панель | Код Доброчесності</title>
<style>
:root{--bg:#f4f7fb;--side:#102f44;--side2:#1f5673;--card:#fff;--ink:#17212f;--mut:#667085;--line:#e7edf3;--gold:#f2b134;--green:#1f8a5b;--red:#b42335;--orange:#b86a00;--shadow:0 18px 50px rgba(16,47,68,.12)}*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--ink)}.app{display:grid;grid-template-columns:270px 1fr;min-height:100vh}.side{background:linear-gradient(180deg,var(--side),var(--side2));color:#fff;padding:22px;position:sticky;top:0;height:100vh}.brand{display:flex;gap:12px;align-items:center;margin-bottom:28px}.logo{width:48px;height:48px;border-radius:16px;background:rgba(255,255,255,.14);display:grid;place-items:center;font-weight:950}.brand b{display:block}.brand span{font-size:12px;color:rgba(255,255,255,.75)}.nav{display:grid;gap:8px}.nav button{width:100%;border:0;background:transparent;color:rgba(255,255,255,.8);text-align:left;padding:12px 14px;border-radius:14px;font-weight:800;cursor:pointer}.nav button.active,.nav button:hover{background:rgba(255,255,255,.14);color:#fff}.main{padding:24px;min-width:0}.top{display:flex;justify-content:space-between;gap:12px;align-items:flex-start;margin-bottom:18px}.top h1{margin:0;color:#102f44}.muted{color:var(--mut);font-size:13px}.card{background:var(--card);border:1px solid #fff;border-radius:22px;box-shadow:var(--shadow);padding:18px;margin-bottom:16px}.grid{display:grid;grid-template-columns:repeat(5,1fr);gap:14px}.stat{padding:18px;border-radius:20px;background:linear-gradient(180deg,#fff,#f8fbfd);border:1px solid var(--line)}.stat b{font-size:28px;color:#102f44}.stat span{display:block;color:var(--mut);font-weight:800;font-size:12px}.toolbar{display:flex;gap:10px;flex-wrap:wrap;align-items:center}.btn{border:0;border-radius:13px;padding:10px 13px;font-weight:850;cursor:pointer;background:#1f5673;color:#fff}.btn.gold{background:var(--gold);color:#102f44}.btn.red{background:var(--red)}.btn.green{background:var(--green)}.btn.orange{background:var(--orange)}.btn.gray{background:#5b6678}input,textarea,select{width:100%;border:1px solid #d7e0e8;border-radius:13px;padding:10px 12px;font-size:14px;background:#fff}label{display:block;font-size:12px;font-weight:850;color:var(--mut);margin:8px 0 5px}.formgrid{display:grid;grid-template-columns:1fr 170px 220px;gap:12px}table{width:100%;border-collapse:collapse}th,td{padding:10px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top;font-size:14px}th{font-size:11px;text-transform:uppercase;color:var(--mut);background:#f7fafc}.pill{display:inline-block;padding:5px 9px;border-radius:999px;background:#e8f1f6;color:#1f5673;font-size:12px;font-weight:850}.pill.soon{background:#fff1d0;color:#8a5200}.pill.today{background:#e2f5ea;color:#18734b}.actions{display:flex;gap:7px;flex-wrap:wrap}.hidden{display:none!important}.status{display:none;margin-bottom:16px;padding:12px 14px;border-radius:15px;font-weight:750}.status.ok{display:block;background:#e7f6ee;color:#18734b}.status.err{display:block;background:#fde8eb;color:#9b1c2b}.modal{position:fixed;inset:0;background:rgba(16,47,68,.42);display:none;place-items:center;padding:18px;z-index:10}.modal.open{display:grid}.modalbox{width:min(860px,100%);max-height:92vh;overflow:auto;background:#fff;border-radius:24px;padding:20px;box-shadow:0 25px 80px rgba(0,0,0,.25)}.chatitem{border:1px solid var(--line);border-radius:18px;padding:14px;margin-bottom:10px}.chatq{font-weight:850;margin:8px 0}.chata{background:#f1f7fb;border-radius:12px;padding:10px;margin:8px 0}.bar{height:12px;background:#edf2f7;border-radius:999px;overflow:hidden}.bar i{display:block;height:100%;background:#1f5673;border-radius:999px}.excelbox{display:flex;gap:10px;flex-wrap:wrap;align-items:end}.excelbox input{max-width:360px}.refitem{border:1px solid var(--line);border-radius:18px;padding:14px;margin-bottom:10px;background:#fff}.refitem b{display:block;color:#102f44;margin-bottom:5px}.pushdiag{white-space:pre-wrap;background:#f7fafc;border:1px solid var(--line);border-radius:14px;padding:12px}@media(max-width:980px){.app{grid-template-columns:1fr}.side{position:relative;height:auto}.grid{grid-template-columns:repeat(2,1fr)}.formgrid{grid-template-columns:1fr}.top{display:block}}@media(max-width:620px){.main{padding:12px}.grid{grid-template-columns:1fr}table,thead,tbody,tr,td,th{display:block}thead{display:none}tr{border:1px solid var(--line);border-radius:14px;margin-bottom:10px;padding:8px;background:#fff}td{border:0}}
</style></head><body><div class="app"><aside class="side"><div class="brand"><div class="logo">КД</div><div><b>Код Доброчесності</b><span>Твій цифровий орієнтир</span></div></div><nav class="nav"><button class="active" data-tab="dashboard">📊 Статистика</button><button data-tab="events">📅 Події</button><button data-tab="reference">📚 Довідник</button><button data-tab="chat">💬 Чат</button><button data-tab="push">🔔 Push</button><button data-tab="settings">⚙️ Налаштування</button></nav></aside><main class="main"><div class="top"><div><h1>Адмін-панель</h1><div class="muted">Керування мобільним додатком</div></div></div><div id="status" class="status"></div>
<section id="dashboard" class="tab"><div class="grid"><div class="stat"><b id="stEvents">0</b><span>Події</span></div><div class="stat"><b id="stViews">0</b><span>Перегляди</span></div><div class="stat"><b id="stCats">0</b><span>Категорії</span></div><div class="stat"><b id="stDevices">0</b><span>Пристрої</span></div><div class="stat"><b id="stChat">0</b><span>Нові питання</span></div></div><div class="card"><div class="toolbar" style="justify-content:space-between"><h2>Найближчі події</h2><button class="btn gray" id="reloadDash">Оновити</button></div><table><thead><tr><th>Дата</th><th>Подія</th><th>Категорія</th><th>Статус</th></tr></thead><tbody id="nearestBody"></tbody></table></div><div class="card"><h2>Перегляди подій</h2><table><thead><tr><th>Подія</th><th>Категорія</th><th>Перегляди</th><th>Графік</th></tr></thead><tbody id="statsBody"></tbody></table></div></section>
<section id="events" class="tab hidden"><div class="card"><div class="toolbar" style="justify-content:space-between"><h2>Події</h2><div class="actions"><button class="btn" id="btnCreate">+ Додати подію</button><label class="btn gold" style="margin:0;cursor:pointer">Імпортувати Excel<input id="excelFile" type="file" accept=".xlsx,.xlsm" style="display:none"></label><button class="btn gold" id="btnExcel">Завантажити</button><button class="btn gray" id="reloadEvents">Оновити</button></div></div><div class="excelbox"><div class="muted">Excel формат: A — дата, B — подія, C — посилання.</div><div id="excelResult" class="muted"></div></div></div><div class="card"><div class="toolbar"><input id="search" placeholder="Пошук події" style="max-width:330px"><select id="filterCat" style="max-width:220px"><option value="">Усі категорії</option><option value="declaration">Декларування</option><option value="conflict">Конфлікт інтересів</option><option value="gifts">Подарунки</option><option value="notice">Повідомлення</option><option value="training">Навчання</option><option value="restriction">Обмеження</option></select><select id="sort" style="max-width:220px"><option value="date">За датою</option><option value="views">За переглядами</option><option value="title">За назвою</option></select></div><table><thead><tr><th>Дата</th><th>Назва</th><th>Категорія</th><th>Перегляди</th><th>Дії</th></tr></thead><tbody id="eventsBody"></tbody></table></div></section>
<section id="reference" class="tab hidden"><div class="card"><h2>Довідник</h2><div class="formgrid"><div><label>Назва</label><input id="refTitle" placeholder="Назва довідкового матеріалу"></div><div><label>Посилання</label><input id="refLink" placeholder="https://..."></div><div style="align-self:end"><button class="btn" id="saveRef">Додати</button></div></div><label>Текст</label><textarea id="refBody" rows="4" placeholder="Короткий опис або інструкція"></textarea></div><div class="card"><div class="toolbar" style="justify-content:space-between"><h2>Матеріали</h2><button class="btn gray" id="reloadRef">Оновити</button></div><div id="refList"></div></div></section>
<section id="chat" class="tab hidden"><div class="card"><div class="toolbar" style="justify-content:space-between"><h2>Чат користувачів</h2><button class="btn gray" id="reloadChat">Оновити</button></div><div id="chatList"></div></div></section>
<section id="push" class="tab hidden"><div class="card"><h2>Push-повідомлення</h2><label>Заголовок</label><input id="pushTitle" placeholder="Наприклад: Нагадування"><label>Текст</label><textarea id="pushBody" rows="4" placeholder="Текст повідомлення"></textarea><div class="actions"><button class="btn green" id="btnPush">Надіслати push</button><button class="btn gray" id="btnPushStatus">Перевірити статус</button></div><div id="pushResult" class="pushdiag muted" style="margin-top:12px"></div></div></section>
<section id="settings" class="tab hidden"><div class="card"><h2>Налаштування</h2><label>ADMIN_TOKEN</label><input id="token" type="password" placeholder="Встав ADMIN_TOKEN"><div class="actions"><button class="btn" id="saveToken">Запам'ятати токен</button><button class="btn gray" id="clearToken">Очистити</button></div></div></section>
</main></div>
<div id="modal" class="modal"><div class="modalbox"><div class="toolbar" style="justify-content:space-between"><h2 id="modalTitle">Нова подія</h2><button class="btn gray" id="closeModal">Закрити</button></div><input type="hidden" id="eventId"><div class="formgrid"><div><label>Назва</label><input id="title"></div><div><label>Дата</label><input id="date" type="date"></div><div><label>Категорія</label><select id="cat"><option value="declaration">Декларування</option><option value="conflict">Конфлікт інтересів</option><option value="gifts">Подарунки</option><option value="notice">Повідомлення</option><option value="training">Навчання</option><option value="restriction">Обмеження</option></select></div></div><label>Повторюваність</label><input id="recur"><label>Опис</label><textarea id="description" rows="3"></textarea><label>Інструкція</label><textarea id="instruction" rows="3"></textarea><label>Посилання</label><input id="link"><label>Аудиторія</label><input id="audience" value="Усі працівники"><label>Нагадування, днів до події</label><input id="reminders" value="30,10,3,0"><div class="actions" style="margin-top:14px"><button class="btn green" id="saveEvent">Зберегти</button></div></div></div>
<script>
const $=id=>document.getElementById(id),catNames={declaration:"Декларування",conflict:"Конфлікт інтересів",gifts:"Подарунки",notice:"Повідомлення",training:"Навчання",restriction:"Обмеження"};let allEvents=[],allChat=[],allRefs=[];$('token').value=sessionStorage.getItem('adminToken')||'';function tok(){return $('token').value.trim()}function h(extra={}){return tok()?{...extra,Authorization:'Bearer '+tok()}:extra}function msg(t,ok=true){const e=$('status');e.className='status '+(ok?'ok':'err');e.textContent=t;e.style.display='block';clearTimeout(e._t);e._t=setTimeout(()=>e.style.display='none',5000)}async function readJson(r){const txt=await r.text();try{return JSON.parse(txt)}catch(e){throw new Error(txt.slice(0,400)||('HTTP '+r.status))}}function fmt(d){if(!d)return'';return new Date(d+'T00:00:00').toLocaleDateString('uk-UA')}function days(d){const a=new Date();a.setHours(0,0,0,0);return Math.ceil((new Date(d+'T00:00:00')-a)/86400000)}function td(t){const c=document.createElement('td');c.textContent=t==null?'':String(t);return c}
document.querySelectorAll('.nav button').forEach(b=>b.onclick=()=>{document.querySelectorAll('.tab').forEach(x=>x.classList.add('hidden'));$(b.dataset.tab).classList.remove('hidden');document.querySelectorAll('.nav button').forEach(x=>x.classList.toggle('active',x===b));if(b.dataset.tab==='dashboard'){loadStats();loadEvents()}if(b.dataset.tab==='chat')loadChat();if(b.dataset.tab==='reference')loadReference();if(b.dataset.tab==='push')pushStatus();});
function openM(edit=false){$('modalTitle').textContent=edit?'Редагування події':'Нова подія';$('modal').classList.add('open')}function closeM(){$('modal').classList.remove('open')}function clearForm(){['eventId','title','date','recur','description','instruction','link'].forEach(id=>$(id).value='');$('cat').value='declaration';$('audience').value='Усі працівники';$('reminders').value='30,10,3,0'}function edit(ev){$('eventId').value=ev.id;$('title').value=ev.title||'';$('date').value=ev.date||'';$('cat').value=ev.cat||'notice';$('recur').value=ev.recur||'';$('description').value=ev.description||'';$('instruction').value=ev.instruction||'';$('link').value=ev.link||'';$('audience').value=ev.audience||'Усі працівники';$('reminders').value=(ev.reminders||[30,10,3,0]).join(',');openM(true)}
function filtered(){let d=[...allEvents],q=$('search').value.toLowerCase().trim(),c=$('filterCat').value;if(c)d=d.filter(e=>e.cat===c);if(q)d=d.filter(e=>[e.title,e.description,e.instruction,e.link,catNames[e.cat]].join(' ').toLowerCase().includes(q));let s=$('sort').value;d.sort((a,b)=>s==='views'?(b.views||0)-(a.views||0):s==='title'?(a.title||'').localeCompare(b.title||'','uk'):(a.date||'').localeCompare(b.date||''));return d}function renderDash(){let views=allEvents.reduce((n,e)=>n+(e.views||0),0),cats=new Set(allEvents.map(e=>e.cat)).size;$('stEvents').textContent=allEvents.length;$('stViews').textContent=views;$('stCats').textContent=cats;let rows=allEvents.map(e=>({...e,d:days(e.date)})).filter(e=>e.d>=0).sort((a,b)=>a.d-b.d).slice(0,8),body=$('nearestBody');body.innerHTML='';if(!rows.length){body.innerHTML='<tr><td colspan="4">Майбутніх подій немає</td></tr>';return}rows.forEach(e=>{let tr=document.createElement('tr');tr.append(td(fmt(e.date)));let name=td(e.title);if(e.link){let m=document.createElement('div');m.className='muted';m.textContent=e.link;name.append(m)}tr.append(name);let cat=document.createElement('td'),p=document.createElement('span');p.className='pill';p.textContent=catNames[e.cat]||e.cat;cat.append(p);tr.append(cat);let st=document.createElement('td'),s=document.createElement('span');s.className='pill '+(e.d===0?'today':e.d<=7?'soon':'');s.textContent=e.d===0?'Сьогодні':'через '+e.d+' дн.';st.append(s);tr.append(st);body.append(tr)})}function renderEvents(){let body=$('eventsBody');body.innerHTML='';let data=filtered();if(!data.length){body.innerHTML='<tr><td colspan="5">Подій не знайдено</td></tr>';return}data.forEach(e=>{let tr=document.createElement('tr');tr.append(td(fmt(e.date)));let name=td(e.title);[e.description,e.link].filter(Boolean).forEach(t=>{let m=document.createElement('div');m.className='muted';m.textContent=t;name.append(m)});tr.append(name);let cat=document.createElement('td'),p=document.createElement('span');p.className='pill';p.textContent=catNames[e.cat]||e.cat;cat.append(p);tr.append(cat);tr.append(td(e.views||0));let a=document.createElement('td'),w=document.createElement('div');w.className='actions';let eb=document.createElement('button');eb.className='btn orange';eb.textContent='Редагувати';eb.onclick=()=>edit(e);let db=document.createElement('button');db.className='btn red';db.textContent='Видалити';db.onclick=()=>del(e.id);w.append(eb,db);a.append(w);tr.append(a);body.append(tr)})}function renderStats(){let body=$('statsBody');body.innerHTML='';let max=Math.max(1,...allEvents.map(e=>e.views||0));[...allEvents].sort((a,b)=>(b.views||0)-(a.views||0)).forEach(e=>{let tr=document.createElement('tr');tr.append(td(e.title));tr.append(td(catNames[e.cat]||e.cat));tr.append(td(e.views||0));let c=document.createElement('td'),bar=document.createElement('div'),i=document.createElement('i');bar.className='bar';i.style.width=Math.max(4,Math.round(((e.views||0)/max)*100))+'%';bar.append(i);c.append(bar);tr.append(c);body.append(tr)})}
async function loadEvents(){try{let r=await fetch('/events');allEvents=await readJson(r);renderDash();renderEvents();renderStats()}catch(e){msg('Помилка подій: '+e.message,false)}}async function loadStats(){if(!tok())return;try{let r=await fetch('/stats',{headers:h()}),s=await readJson(r);$('stChat').textContent=s.chat_new||0;$('stDevices').textContent=s.devices||0;if(s.ok===false)msg('Статистика: '+s.error,false)}catch(e){msg('Помилка статистики: '+e.message,false)}}
async function save(){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}let id=$('eventId').value.trim(),payload={title:$('title').value.trim(),date:$('date').value,cat:$('cat').value,recur:$('recur').value.trim(),description:$('description').value.trim(),instruction:$('instruction').value.trim(),link:$('link').value.trim(),audience:$('audience').value.trim()||'Усі працівники',reminders:$('reminders').value.split(',').map(x=>Number(x.trim())).filter(x=>!isNaN(x))};if(!payload.title||!payload.date){msg('Заповни назву і дату',false);return}try{let r=await fetch(id?`/events/${id}`:'/events',{method:id?'PUT':'POST',headers:h({'Content-Type':'application/json'}),body:JSON.stringify(payload)});if(!r.ok)throw new Error((await readJson(r)).detail||'Помилка');closeM();clearForm();await loadEvents();await loadStats();msg('Збережено')}catch(e){msg('Помилка: '+e.message,false)}}async function del(id){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}if(!confirm('Видалити подію?'))return;let r=await fetch(`/events/${id}`,{method:'DELETE',headers:h()});if(r.ok||r.status===204){await loadEvents();await loadStats();msg('Видалено')}else msg('Помилка видалення',false)}async function uploadExcel(){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}let f=$('excelFile').files[0];if(!f){msg('Вибери Excel файл',false);return}let fd=new FormData();fd.append('file',f);try{let r=await fetch('/events/import-excel',{method:'POST',headers:h(),body:fd});let j=await readJson(r);if(!r.ok)throw new Error(j.detail||JSON.stringify(j));$('excelResult').textContent=`Завантажено: ${j.created}. Пропущено: ${j.skipped_count}.`;await loadEvents();await loadStats();msg('Excel завантажено')}catch(e){msg('Помилка Excel: '+e.message,false)}}
async function loadReference(){try{let r=await fetch('/reference');allRefs=await readJson(r);let box=$('refList');box.innerHTML='';if(!allRefs.length){box.innerHTML='<div class="muted">Матеріалів поки немає</div>';return}allRefs.forEach(x=>{let d=document.createElement('div');d.className='refitem';let b=document.createElement('b');b.textContent=x.title;let p=document.createElement('div');p.textContent=x.body||'';let l=document.createElement('div');l.className='muted';l.textContent=x.link||'';let a=document.createElement('div');a.className='actions';let delb=document.createElement('button');delb.className='btn red';delb.textContent='Видалити';delb.onclick=()=>deleteRef(x.id);a.append(delb);d.append(b,p,l,a);box.append(d)})}catch(e){msg('Помилка довідника: '+e.message,false)}}async function saveRef(){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}let payload={title:$('refTitle').value.trim(),body:$('refBody').value.trim(),link:$('refLink').value.trim()};if(!payload.title){msg('Вкажи назву довідника',false);return}let r=await fetch('/reference',{method:'POST',headers:h({'Content-Type':'application/json'}),body:JSON.stringify(payload)});if(r.ok){$('refTitle').value='';$('refBody').value='';$('refLink').value='';loadReference();loadStats();msg('Додано в довідник')}else msg('Помилка довідника',false)}async function deleteRef(id){if(!confirm('Видалити запис?'))return;let r=await fetch(`/reference/${id}`,{method:'DELETE',headers:h()});if(r.ok||r.status===204){loadReference();loadStats();msg('Запис видалено')}else msg('Помилка видалення',false)}
async function loadChat(){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}try{let r=await fetch('/admin/chat',{headers:h()});allChat=await readJson(r);let box=$('chatList');box.innerHTML='';if(!allChat.length){box.innerHTML='<div class="muted">Повідомлень поки немає</div>';return}allChat.forEach(m=>{let d=document.createElement('div');d.className='chatitem';d.innerHTML='<div class="muted"></div><div class="chatq"></div><textarea rows="3" placeholder="Відповідь"></textarea><div class="actions"><button class="btn green">Відповісти</button><button class="btn red">Видалити</button></div>';d.querySelector('.muted').textContent=m.client_id+' • '+new Date(m.created_at).toLocaleString('uk-UA');d.querySelector('.chatq').textContent=m.question;if(m.answer){let a=document.createElement('div');a.className='chata';a.textContent='Відповідь: '+m.answer;d.insertBefore(a,d.querySelector('textarea'));d.querySelector('textarea').value=m.answer}d.querySelector('.green').onclick=()=>answerChat(m.id,d.querySelector('textarea').value);d.querySelector('.red').onclick=()=>deleteChat(m.id);box.append(d)})}catch(e){msg('Помилка чату: '+e.message,false)}}async function answerChat(id,answer){let r=await fetch(`/admin/chat/${id}/answer`,{method:'POST',headers:h({'Content-Type':'application/json'}),body:JSON.stringify({answer})});if(r.ok){msg('Відповідь збережено');loadChat();loadStats()}else msg('Помилка відповіді',false)}async function deleteChat(id){if(!confirm('Видалити повідомлення?'))return;let r=await fetch(`/admin/chat/${id}`,{method:'DELETE',headers:h()});if(r.ok||r.status===204){msg('Чат видалено');loadChat();loadStats()}else msg('Помилка видалення',false)}
async function sendPush(){if(!tok()){msg('Спершу вкажи ADMIN_TOKEN',false);return}let title=$('pushTitle').value.trim(),body=$('pushBody').value.trim();if(!title||!body){msg('Заповни заголовок і текст',false);return}try{let r=await fetch('/push/send',{method:'POST',headers:h({'Content-Type':'application/json'}),body:JSON.stringify({title,body})});let j=await readJson(r);$('pushResult').textContent=JSON.stringify(j,null,2);msg(j.configured&&j.sent>0?'Push відправлено':'Push не відправлено: '+(j.error||'дивись статус'),j.configured&&j.sent>0)}catch(e){msg('Помилка push: '+e.message,false)}}async function pushStatus(){if(!tok())return;try{let r=await fetch('/push/status',{headers:h()}),j=await readJson(r);$('pushResult').textContent=JSON.stringify(j,null,2)}catch(e){$('pushResult').textContent='Помилка статусу: '+e.message}}
$('btnCreate').onclick=()=>{clearForm();openM(false)};$('closeModal').onclick=closeM;$('modal').onclick=e=>{if(e.target.id==='modal')closeM()};$('saveEvent').onclick=save;$('reloadEvents').onclick=$('reloadDash').onclick=()=>{loadEvents();loadStats()};['search','filterCat','sort'].forEach(id=>$(id).oninput=renderEvents);$('saveToken').onclick=()=>{sessionStorage.setItem('adminToken',tok());msg('Токен збережено');loadStats();pushStatus()};$('clearToken').onclick=()=>{sessionStorage.removeItem('adminToken');$('token').value='';msg('Токен очищено')};$('btnExcel').onclick=uploadExcel;$('reloadChat').onclick=loadChat;$('saveRef').onclick=saveRef;$('reloadRef').onclick=loadReference;$('btnPush').onclick=sendPush;$('btnPushStatus').onclick=pushStatus;loadEvents();loadStats();loadReference();
</script></body></html>
'''

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main_fixed_admin:app", host="0.0.0.0", port=8000, reload=True)
